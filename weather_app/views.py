import csv
import zipfile
from StringIO import StringIO
from io import BytesIO

from django.conf import settings
from django.core.cache import cache
from django.http import HttpResponse
from django.views.decorators.csrf import csrf_exempt
from openpyxl import load_workbook

from .models import GeoPosition, WeatherEntry, AverageStats


@csrf_exempt
def upload(request):
    if not request.method == 'POST':
        return HttpResponse()

    wb = load_workbook(filename=BytesIO(request.body))
    lines_gen = wb[wb.sheetnames[0]].values
    titles = lines_gen.next()
    entries = []

    for line in lines_gen:
        entry = dict(zip(titles, line))
        entries.append(entry)

    for entry in entries:
        geoposition, _ = GeoPosition.objects.get_or_create(
            coordinates=entry['geoposition coordinates'],
            height=entry['hight under the ground']
        )

        WeatherEntry.objects.create(
            geoposition=geoposition,
            date=entry['date'],
            time=entry['time'],
            temperature=entry['temperature'],
            wind_speed=entry['wind speed'],
            wind_vector=entry['wind vector direction']
        )

    return HttpResponse('Successfully uploaded %s entries!' % len(entries))


def calculate_statistics():
    geopositions = GeoPosition.objects.all()
    stats = [
        AverageStats.objects.filter(geoposition=geoposition).latest('datetime')
        for geoposition in geopositions
    ]
    payload = [('Coordinates', 'Height', 'Average temperature',
                'Average wind speed', 'Average wind vector', 'Last update')]
    payload.extend([
        (stat.geoposition.coordinates,
         stat.geoposition.height,
         stat.avg_temperature,
         stat.avg_wind_speed,
         stat.avg_wind_vector,
         stat.datetime.strftime("%I:%M%p on %B %d, %Y")) for stat in stats
    ])
    cache.set(settings.UPDATE_WEATHER_STATISTICS_CACHE_KEY, payload,
              timeout=settings.UPDATE_WEATHER_STATISTICS_TASK_PERIOD)

    return payload


def statistics(request):
    payload = cache.get(settings.UPDATE_WEATHER_STATISTICS_CACHE_KEY)
    if payload is None:
        payload = calculate_statistics()

    s_csv = StringIO()
    s_zip = BytesIO()

    csv.writer(s_csv, dialect='excel').writerows(payload)

    with zipfile.ZipFile(s_zip, 'w', zipfile.ZIP_DEFLATED) as zip:
        zip.writestr('weather-statistics.csv', s_csv.getvalue())

    resp = HttpResponse(
        s_zip.getvalue(),
        content_type='application/x-zip-compressed'
    )
    resp['Content-Disposition'] = 'attachment; filename=weather-statistics.zip'

    return resp
